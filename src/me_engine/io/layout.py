"""Geometric layout of an ME geography sheet — discovered, not hard-coded.

The source workbook is regular but has small per-sheet variations (country sheets
omit the 'By Country' block, so the volume band's rows shift). Rather than encode
fragile row offsets, we *discover* the layout by scanning a sheet once into a
`SheetLayout`: band title rows + a (band, label) -> row map. Reader and writer
both address cells through this same discovered map, guaranteeing they agree.
"""
from __future__ import annotations

from dataclasses import dataclass
from functools import cached_property
from typing import Mapping

from openpyxl.worksheet.worksheet import Worksheet

from ..domain.taxonomy import Band, YEARS

# --- Column geometry (1-based) ----------------------------------------------
COL_LABEL = 3                  # C: row label
COL_METRIC_START = 4           # D: first year (2021)
COL_YOY_START = 18             # R: first Y-o-Y year (2022)
COL_CAGR = 31                  # AE
COL_SHARE_START = 37           # AK: first share year (2021)

_BAND_BY_TITLE = {b.value: b for b in Band}


def metric_col(year: int) -> int:
    return COL_METRIC_START + YEARS.index(year)


def yoy_col(year: int) -> int:
    return COL_YOY_START + (year - YEARS[1])


def share_col(year: int) -> int:
    return COL_SHARE_START + YEARS.index(year)


def cagr_col_for() -> int:
    return COL_CAGR


@dataclass(frozen=True, slots=True)
class SheetLayout:
    """Discovered positions for one geography sheet."""

    band_title_row: Mapping[Band, int]
    row_of: Mapping[tuple[Band, str], int]      # (band, label) -> row number

    @classmethod
    def discover(cls, ws: Worksheet) -> "SheetLayout":
        """Scan a worksheet's label column to locate every band and data row.

        Band membership is determined by which band-title we have most recently
        passed, so each label is filed under its correct band even when the same
        label (e.g. 'Bottles') appears in Value, ASP, and Volume.
        """
        titles: dict[Band, int] = {}
        rows: dict[tuple[Band, str], int] = {}
        current: Band | None = None

        for r in range(1, ws.max_row + 1):
            marker = ws.cell(r, COL_METRIC_START).value          # band title (col D)
            if isinstance(marker, str) and marker in _BAND_BY_TITLE:
                current = _BAND_BY_TITLE[marker]
                titles[current] = r
                continue
            label = ws.cell(r, COL_LABEL).value
            if current is not None and isinstance(label, str) and label.strip():
                rows[(current, label)] = r

        return cls(band_title_row=titles, row_of=rows)

    def labels_in(self, band: Band) -> tuple[str, ...]:
        return tuple(label for (b, label) in self.row_of if b == band)
