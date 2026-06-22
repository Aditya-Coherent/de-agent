"""Render a MarketResult into an ME workbook in the exact source style.

Strategy: load the reference workbook as a *style template* and overwrite only
the numeric cells (metric paths, Y-o-Y, CAGR, market share) at positions found by
`SheetLayout`. Because we never touch formatting, headers or spacing, the output
is style-identical to the source; only the numbers are (re)written from the
assembled result. When no template is supplied, a blank workbook could be built,
but template-based rendering is what guarantees parity for Stage 1.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..assembly.model import BandResult, MarketResult, MetricRow
from ..domain.taxonomy import YEARS
from .layout import SheetLayout, cagr_col_for, metric_col, share_col, yoy_col


class MEWorkbookWriter:
    """Writes a MarketResult onto a copy of a reference template workbook."""

    def __init__(self, template_path: Path | str) -> None:
        self._template_path = Path(template_path)

    def write(self, result: MarketResult, out_path: Path | str) -> Path:
        wb = load_workbook(self._template_path)        # keeps styles & layout
        for name, geo in result.geographies.items():
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            layout = SheetLayout.discover(ws)
            for band_result in geo.bands.values():
                self._write_band(ws, layout, band_result)
        out = Path(out_path)
        wb.save(out)
        return out

    def _write_band(self, ws: Worksheet, layout: SheetLayout, band: BandResult) -> None:
        for label, row in band.rows_by_label.items():
            coord = layout.row_of.get((band.band, label))
            if coord is not None:
                self._write_row(ws, coord, row)

    def _write_row(self, ws: Worksheet, coord: int, row: MetricRow) -> None:
        self._write_metric(ws, coord, row)
        self._write_yoy(ws, coord, row)
        self._write_cagr(ws, coord, row)
        self._write_share(ws, coord, row)

    @staticmethod
    def _write_metric(ws: Worksheet, coord: int, row: MetricRow) -> None:
        for year in YEARS:
            ws.cell(coord, metric_col(year)).value = row.series.at(year)

    @staticmethod
    def _write_yoy(ws: Worksheet, coord: int, row: MetricRow) -> None:
        for year, value in row.series.yoy().items():
            ws.cell(coord, yoy_col(year)).value = value

    @staticmethod
    def _write_cagr(ws: Worksheet, coord: int, row: MetricRow) -> None:
        ws.cell(coord, cagr_col_for()).value = row.series.cagr()

    @staticmethod
    def _write_share(ws: Worksheet, coord: int, row: MetricRow) -> None:
        if row.share_of_parent is None:
            return
        for year in YEARS:
            ws.cell(coord, share_col(year)).value = row.share_of_parent.at(year)
