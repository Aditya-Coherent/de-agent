"""Read the Input Sheet into per-geography curve context.

The Input Sheet is the agent's starting point: it supplies the forecast CAGR, the
2025 anchor value, and the region each geography belongs to. (It does NOT contain
the growth curve — that is exactly what the Curve Agent produces.)
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ..curve.agent import MarketContext
from ..domain.taxonomy import GEOGRAPHIES

_NAME_COL, _CAGR_COL, _VALUE_COL = 2, 4, 5      # cols B, D, E
_MARKET_NAME_CELL = (1, 3)                      # C1


@dataclass(frozen=True, slots=True)
class GeoInput:
    name: str
    forecast_cagr: float
    anchor_2025: float
    region: str | None


class InputSheetReader:
    """Extracts geography CAGR/value/region rows from the Input Sheet."""

    def __init__(self, path: Path | str, sheet: str = "Data") -> None:
        self._ws = load_workbook(Path(path), data_only=True)[sheet]

    def market_name(self) -> str:
        return str(self._ws.cell(*_MARKET_NAME_CELL).value)

    def read(self) -> dict[str, GeoInput]:
        """Geography rows from the sizing block, keyed by name.

        Only rows whose label is a *known geography* are kept — this excludes the
        segmentation rows (product/packaging/etc.) at the top of the sheet that
        also carry numbers. The sizing block lists each geography with its CAGR
        and 2025 value; the last occurrence wins (dedupes the two sub-blocks).
        """
        valid_names = set(GEOGRAPHIES.by_name)
        parent_of = GEOGRAPHIES.parent_of
        found: dict[str, GeoInput] = {}
        for row in range(1, self._ws.max_row + 1):
            name = self._ws.cell(row, _NAME_COL).value
            cagr = self._ws.cell(row, _CAGR_COL).value
            value = self._ws.cell(row, _VALUE_COL).value
            if (isinstance(name, str) and name in valid_names
                    and isinstance(cagr, (int, float))
                    and isinstance(value, (int, float))):
                found[name] = GeoInput(
                    name=name,
                    forecast_cagr=float(cagr),
                    anchor_2025=float(value),
                    region=parent_of.get(name),
                )
        return found

    def contexts(self) -> list[MarketContext]:
        market = self.market_name()
        return [
            MarketContext(market_name=market, geography=g.name,
                          forecast_cagr=g.forecast_cagr, region=g.region)
            for g in self.read().values()
        ]
