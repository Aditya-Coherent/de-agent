"""Build a DriverSet from the Input Sheet + agent curves (the pilot data path).

This is the live-system analogue of MEWorkbookReader: instead of reading finished
dynamics out of an ME file, it composes them from
  - the Input Sheet's single-year snapshots (segmentation %, ASP), and
  - the Curve Agent's value path per geography.

Where the Input Sheet has only a single year, we hold that value flat across the
horizon. This is an explicit, measurable assumption: the gap it creates versus the
human ME file quantifies exactly how much the (not-yet-built) segmentation/ASP
drift agents still need to supply.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..curve.asp_agent import AspAgent
from ..curve.runner import CurveRunner, truth_paths_from_me
from ..curve.seg_agent import SegmentationAgent
from ..curve.segmentation import ShareDriftBuilder
from ..domain.drivers import (
    AspDrivers, DriverSet, GeographyDrivers, SegmentationDrivers,
)
from ..domain.series import Series
from ..domain.taxonomy import (
    BASE_YEAR, GEOGRAPHIES, PRICED_DIMENSION, SEGMENTATION_DIMENSIONS, YEARS,
)
from .input_reader import InputSheetReader


def _flat(value: float) -> Series:
    """Hold a single-year snapshot flat across the whole horizon."""
    return Series({year: value for year in YEARS})


class InputDriverBuilder:
    """Composes Input-Sheet snapshots + agent curves into a DriverSet."""

    def __init__(self, input_path: Path | str,
                 truth_me: Path | str | None = None,
                 use_segmentation_agent: bool = True) -> None:
        self._input_path = Path(input_path)
        self._data = load_workbook(self._input_path, data_only=True)["Data"]
        self._asp = load_workbook(self._input_path, data_only=True)["ASP"]
        self._truth = truth_paths_from_me(truth_me) if truth_me else None
        self._col_of = self._scan_country_columns()
        self._drift = ShareDriftBuilder()
        market = InputSheetReader(self._input_path).market_name()
        self._premiums = (self._decide_premiums(market)
                          if use_segmentation_agent else {})
        self._asp_agent = AspAgent()
        self._asp_decision = self._asp_agent.decide(market) if use_segmentation_agent else None

    def _decide_premiums(self, market: str) -> dict:
        """Decide per-segment growth premiums once per market (structural)."""
        agent = SegmentationAgent()
        return {dim.title: agent.decide(market, dim).premiums
                for dim in SEGMENTATION_DIMENSIONS}

    def build(self) -> DriverSet:
        reader = InputSheetReader(self._input_path)
        inputs = reader.read()
        self._curve_outcomes = {
            o.geo.name: o for o in CurveRunner().run(self._input_path, self._truth)
        }

        geographies = {}
        for name, geo in inputs.items():
            if name not in self._col_of:
                continue                          # only geos with a segmentation column
            value = (self._curve_outcomes[name].path
                     if name in self._curve_outcomes else _flat(geo.anchor_2025))
            geographies[name] = GeographyDrivers(
                name=name,
                value=value,
                segmentation=self._segmentation(name, geo.forecast_cagr),
                asp=self._asp_drivers(name),
            )
        return DriverSet(market_name=reader.market_name(), geographies=geographies)

    def curve_rationale(self) -> dict[str, str]:
        """Return per-geography curve agent reasoning after build() is called."""
        if not hasattr(self, "_curve_outcomes"):
            return {}
        rationale = {}
        for name, outcome in self._curve_outcomes.items():
            d = outcome.decision
            tag = "FALLBACK" if d.used_fallback else "AGENT"
            score_str = ""
            if outcome.score:
                sc = outcome.score
                score_str = (f" | shape_err={sc.shape_mae:.4f} "
                             f"cosine={sc.cosine:.4f} path_max={sc.path_max_rel_error:.2%}")
            rationale[name] = (
                f"[{tag}] archetype={d.archetype} peak={d.peak_year} "
                f"conf={d.confidence:.2f}{score_str} | {d.reasoning}"
            )
        return rationale

    # --- segmentation & asp from single-year snapshots ----------------------
    def _segmentation(self, geo: str, market_cagr: float) -> SegmentationDrivers:
        """Base-year shares from the Input Sheet, drifted by agent premiums.

        When no premiums are available the drift is zero and shares stay flat,
        recovering the earlier baseline behaviour exactly.
        """
        col = self._col_of[geo]
        shares: dict[str, dict[str, Series]] = {}
        for dim in SEGMENTATION_DIMENSIONS:
            base = self._read_base_shares(dim, col)
            premiums = self._premiums.get(dim.title, {})
            if premiums:
                shares[dim.title] = self._drift.build(dim, base, market_cagr, premiums)
            else:
                shares[dim.title] = {seg: _flat(v) for seg, v in base.items()}
        return SegmentationDrivers(shares=shares)

    def _asp_drivers(self, geo: str) -> AspDrivers:
        """Base-year ASP per product, grown by the agent's inflation rate.

        With no ASP decision (segmentation agent disabled) prices stay flat,
        recovering the earlier baseline.
        """
        col = self._col_of[geo]
        asp: dict[str, Series] = {}
        for product in PRICED_DIMENSION.segments:
            base = self._asp_cell(self._asp_row(product), col)
            if self._asp_decision is not None:
                rate = self._asp_decision.rates[product]
                asp[product] = self._asp_agent.price_path(base, rate)
            else:
                asp[product] = _flat(base)
        return AspDrivers(asp=asp)

    def _read_base_shares(self, dim, col: int) -> dict[str, float]:
        """Read base-year shares for one dimension, inferring None cells.

        Some input sheets store only N-1 of N siblings — the last/first is
        left blank and implied as the complement (1 - sum of siblings with the
        same parent). This is common for the Distribution Channel hierarchy.
        """
        raw: dict[str, float | None] = {}
        for seg in dim.segments:
            v = self._data.cell(self._segment_row(dim, seg), col).value
            raw[seg] = float(v) if isinstance(v, (int, float)) else None

        # Group siblings by parent and infer any single None as complement.
        from collections import defaultdict
        by_parent: dict[str | None, list[str]] = defaultdict(list)
        for seg in dim.segments:
            by_parent[dim.parent_of(seg)].append(seg)

        result: dict[str, float] = {}
        for parent, siblings in by_parent.items():
            nones = [s for s in siblings if raw[s] is None]
            known = [s for s in siblings if raw[s] is not None]
            if len(nones) == 0:
                for s in siblings:
                    result[s] = raw[s]
            elif len(nones) == 1:
                # Infer missing sibling as complement of the rest
                known_sum = sum(raw[s] for s in known)
                result[nones[0]] = max(0.0, 1.0 - known_sum)
                for s in known:
                    result[s] = raw[s]
            else:
                # Multiple Nones: fall back to equal split of remainder
                known_sum = sum(raw[s] for s in known) if known else 0.0
                remainder = max(0.0, 1.0 - known_sum)
                share_each = remainder / len(nones)
                for s in known:
                    result[s] = raw[s]
                for s in nones:
                    result[s] = share_each

        return result

    # --- sheet scanning helpers --------------------------------------------
    # The segmentation %% band occupies the first country header band only; the
    # sheet repeats country headers further right for CAGR and trend-ID bands.
    _SEGMENTATION_BAND_END_COL = 30      # column AD (last country of the % band)

    def _scan_country_columns(self) -> dict[str, int]:
        """Map geography name -> its column in the first segmentation band (C..AD).

        Only the first band is scanned; later repeats (CAGR band, CMI trend-ID
        band) reuse the same country headers and would otherwise overwrite the
        mapping with non-percentage columns.
        """
        valid = set(GEOGRAPHIES.by_name)
        return {
            self._data.cell(3, c).value: c
            for c in range(3, self._SEGMENTATION_BAND_END_COL + 1)
            if self._data.cell(3, c).value in valid
        }

    def _segment_row(self, dim, segment: str) -> int:
        return self._find_row(self._data, segment)

    def _asp_row(self, product: str) -> int:
        return self._find_row(self._asp, product)

    @staticmethod
    def _find_row(ws, label: str) -> int:
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 2).value == label:
                return r
        raise KeyError(f"label not found: {label}")

    def _cell(self, row: int, col: int) -> float:
        return float(self._data.cell(row, col).value)

    def _asp_cell(self, row: int, col: int) -> float:
        return float(self._asp.cell(row, col).value)
