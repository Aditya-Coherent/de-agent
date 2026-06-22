"""Read an existing ME workbook into a typed DriverSet.

This is the Stage-1 data source: it recovers the full year-by-year dynamics
(value paths, segment-share drift, ASP paths) that the Input Sheet does not yet
contain, so the assembler has a complete driver set to reproduce against. In the
live system an agent layer replaces this reader; nothing downstream changes.
"""
from __future__ import annotations

from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..domain.drivers import (
    AspDrivers, DriverSet, GeographyDrivers, SegmentationDrivers,
)
from ..domain.series import Series
from ..domain.taxonomy import (
    Band, GEOGRAPHIES, PRICED_DIMENSION, SEGMENTATION_DIMENSIONS, YEARS,
)
from .layout import SheetLayout, metric_col


class MEWorkbookReader:
    """Loads a workbook once and exposes a `DriverSet` view of it."""

    def __init__(self, path: Path | str) -> None:
        self._path = Path(path)
        self._wb = load_workbook(self._path, data_only=True, read_only=True)

    def read(self, market_name: str) -> DriverSet:
        geographies = {
            geo.name: self._read_geography(geo.name)
            for geo in GEOGRAPHIES.in_order
            if geo.name in self._wb.sheetnames
        }
        return DriverSet(market_name=market_name, geographies=geographies)

    # --- per-geography extraction ------------------------------------------
    def _read_geography(self, name: str) -> GeographyDrivers:
        ws = self._wb[name]
        layout = SheetLayout.discover(ws)
        value = self._value_series(ws, layout)
        return GeographyDrivers(
            name=name,
            value=value,
            segmentation=self._segmentation(ws, layout, value),
            asp=self._asp(ws, layout),
        )

    def _value_series(self, ws: Worksheet, layout: SheetLayout) -> Series:
        """The band total for Value (taken from the first segmentation header)."""
        header_label = SEGMENTATION_DIMENSIONS[0].title
        row = layout.row_of[(Band.VALUE, header_label)]
        return self._row_series(ws, row)

    def _segmentation(
        self, ws: Worksheet, layout: SheetLayout, total: Series,
    ) -> SegmentationDrivers:
        """Recover each segment's *share path*, relative to its immediate parent.

        Flat dimensions divide by the band total; hierarchical channels divide by
        the parent segment's value — matching how the workbook stores shares.
        """
        shares: dict[str, dict[str, Series]] = {}
        for dim in SEGMENTATION_DIMENSIONS:
            values = {
                segment: self._row_series(ws, layout.row_of[(Band.VALUE, segment)])
                for segment in dim.segments
            }
            shares[dim.title] = {
                segment: series.share_of(
                    total if dim.parent_of(segment) is None
                    else values[dim.parent_of(segment)])
                for segment, series in values.items()
            }
        return SegmentationDrivers(shares=shares)

    def _asp(self, ws: Worksheet, layout: SheetLayout) -> AspDrivers:
        return AspDrivers(asp={
            product: self._row_series(ws, layout.row_of[(Band.ASP, product)])
            for product in PRICED_DIMENSION.segments
        })

    @staticmethod
    def _row_series(ws: Worksheet, row: int) -> Series:
        return Series({year: float(ws.cell(row, metric_col(year)).value)
                       for year in YEARS})
