"""Cell-for-cell verification of an assembled MarketResult against a source ME file.

The diff reads the metric cells of every band/row/year from the source workbook
and compares them to the assembled values, reporting the worst and average
absolute deviations. This is the Stage-1 definition of done: deviation ~ 0.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from ..assembly.model import MarketResult
from ..domain.taxonomy import Band, YEARS
from ..io.layout import SheetLayout, metric_col


@dataclass(frozen=True, slots=True)
class Deviation:
    sheet: str
    band: Band
    label: str
    year: int
    expected: float
    actual: float

    @property
    def abs_error(self) -> float:
        return abs(self.expected - self.actual)

    @property
    def rel_error(self) -> float:
        return self.abs_error / abs(self.expected) if self.expected else self.abs_error


@dataclass(frozen=True, slots=True)
class DiffReport:
    deviations: tuple[Deviation, ...]
    compared: int

    @property
    def max_abs(self) -> float:
        return max((d.abs_error for d in self.deviations), default=0.0)

    @property
    def max_rel(self) -> float:
        return max((d.rel_error for d in self.deviations), default=0.0)

    def worst(self, n: int = 10) -> tuple[Deviation, ...]:
        return tuple(sorted(self.deviations, key=lambda d: d.abs_error, reverse=True)[:n])

    def passed(self, tol: float = 1e-6) -> bool:
        return self.max_abs <= tol


class DiffEngine:
    """Compares an assembled result to a reference workbook."""

    def __init__(self, reference_path: Path | str, tol: float = 1e-6) -> None:
        self._wb = load_workbook(Path(reference_path), data_only=True, read_only=True)
        self._tol = tol

    def compare(self, result: MarketResult) -> DiffReport:
        deviations = tuple(self._scan(result))
        compared = self._count(result)
        return DiffReport(deviations=deviations, compared=compared)

    def _scan(self, result: MarketResult) -> Iterator[Deviation]:
        for name, geo in result.geographies.items():
            if name not in self._wb.sheetnames:
                continue
            ws = self._wb[name]
            layout = SheetLayout.discover(ws)
            for band, band_result in geo.bands.items():
                for label, row in band_result.rows_by_label.items():
                    coord = layout.row_of.get((band, label))
                    if coord is None:
                        continue
                    yield from self._scan_row(name, band, label, coord, row, ws)

    def _scan_row(self, sheet, band, label, coord, row, ws) -> Iterator[Deviation]:
        for year in YEARS:
            cell = ws.cell(coord, metric_col(year)).value
            if cell is None:
                continue
            expected = float(cell)
            actual = row.series.at(year)
            if abs(expected - actual) > self._tol:
                yield Deviation(sheet, band, label, year, expected, actual)

    def _count(self, result: MarketResult) -> int:
        return sum(
            len(band.rows_by_label) * len(YEARS)
            for geo in result.geographies.values()
            for band in geo.bands.values()
        )
